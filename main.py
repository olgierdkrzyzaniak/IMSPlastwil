import tkinter as tk
from tkinter import ttk
from datetime import datetime
import openpyxl as xl

CONSTANTS = {
    "take product": "1",
    "return product": "2",
    "cancel": "3"
}


class InventoryDatabase:
    def __init__(self, excel_file):
        self.excel_file = excel_file
        self.workbook = None
        self.products = None
        self.users = None
        self.activities = None

        self.load_inventory()

    def load_inventory(self):
        try:
            self.workbook = xl.load_workbook(self.excel_file)
            self.products = self.workbook['Products']
            self.users = self.workbook['Users']
            self.activities = self.workbook['Activities']
        except FileNotFoundError:
            self.workbook = xl.Workbook()
            self.products = self.workbook.create_sheet("Products")
            self.products.append(["Kod produktu", "Product Name", "Ilość"])
            self.users = self.workbook.create_sheet("Users")
            self.users.append(["Kod pracownika", "Username"])
            self.activities = self.workbook.create_sheet("Activities")
            self.activities.append(["Kod pracownika", "Kod produktu", "Ilość", "Czynność", "Data"])
            self.workbook.save(self.excel_file)

    def find_user_by_code(self, user_code):
        for row in self.users.iter_rows(min_row=2, values_only=True):
            if str(row[0]) == str(user_code):
                return row
        return None

    def find_product_by_code(self, product_code):
        for row in self.products.iter_rows(min_row=2, values_only=False):
            if str(row[0].value) == str(product_code):
                return row
        return None

    def update_current_user(self, user_code):
        user = self.find_user_by_code(user_code)
        if user:
            self.current_user = user[1]
        else:
            self.current_user = None

    def add_activity(self, user_id, product_id, quantity, activity, date):
        self.activities.append([user_id, product_id, quantity, activity, date])
        self.workbook.save(self.excel_file)


class InventoryApp:
    def __init__(self, root, database):
        self.root = root
        self.root.title('VOESTALPINE Inventory Management System')

        self.database = database
        self.current_user = None
        self.record_id = 1

        self.setup_ui()

    def setup_ui(self):
        self.create_widgets()
        self.arrange_widgets()

    def create_widgets(self):
        # Radio buttons
        self.activity_var = tk.StringVar(value="Take Product")
        self.radio_take = ttk.Radiobutton(self.root, text="Pobierz Produkt", variable=self.activity_var,
                                          value="Take Product")
        self.radio_return = ttk.Radiobutton(self.root, text="Zwróć Produkt", variable=self.activity_var,
                                            value="Return Product")

        # Current user label
        self.label_current_user = ttk.Label(self.root, text="Aktualny Użytkownik: None", foreground="red")

        # Error label
        self.error_label = ttk.Label(self.root, text="", foreground="red")

        # Barcode input
        self.label_barcode = ttk.Label(self.root, text="Kod kreskowy:")
        self.style = ttk.Style()
        self.style.configure('TEntry', width=20)  # Explicitly set width
        self.entry_barcode = ttk.Entry(self.root, style='TEntry')
        self.submit_button = ttk.Button(self.root, text="Dodaj", command=self.submit_action)

        # Table
        self.table_columns = ("ID", "Kod pracownika", "Kod produktu", "Ilość", "Czynność", "Data")
        self.tree = ttk.Treeview(self.root, columns=self.table_columns, show="headings", selectmode="browse")

        # Set column headings
        for col in self.table_columns:
            self.tree.heading(col, text=col)

        # Save button
        self.save_button = ttk.Button(self.root, text="Zapisz", command=self.save_data)

        #Cancel button
        self.cancel_button = ttk.Button(self.root, text="Anuluj", command=self.cancel_action)

    def arrange_widgets(self):
        # Grid layout
        self.radio_take.grid(row=0, column=0, sticky="w", padx=5, pady=5)
        self.radio_return.grid(row=1, column=0, sticky="w", padx=5, pady=5)
        self.label_current_user.grid(row=2, column=0, sticky="w", padx=5, pady=5)
        self.error_label.grid(row=2, column=1, columnspan=2, sticky="w", padx=5, pady=5)
        self.label_barcode.grid(row=3, column=0, sticky="w", padx=5, pady=5)
        self.entry_barcode.grid(row=3, column=1, sticky="w", padx=5, pady=5)
        self.submit_button.grid(row=3, column=2, sticky="w", padx=5, pady=5)
        self.tree.grid(row=4, column=0, columnspan=3, sticky="w", padx=5, pady=5)
        self.cancel_button.grid(row=5, column=0, sticky="w", padx=5, pady=5)  # Align to the right
        self.save_button.grid(row=5, column=0, columnspan=3, sticky="e", padx=5, pady=5)  # Align to the right

        # Bind Enter key to submit action
        self.entry_barcode.bind('<Return>', lambda event: self.submit_action())

        # Set default focus on barcode input
        self.entry_barcode.focus_set()

    def cancel_action(self):
        # Clear the table
        for item_id in self.tree.get_children():
            self.tree.delete(item_id)

        # Clear the input field
        self.entry_barcode.delete(0, tk.END)

        # Set default focus on barcode input
        self.entry_barcode.focus_set()

    def submit_action(self):
        #clear error label
        self.error_label.config(text="")
        barcode = self.entry_barcode.get()
        if barcode:
            if barcode == CONSTANTS["take product"]:
                self.activity_var.set("Take Product")
            elif barcode == CONSTANTS["return product"]:
                self.activity_var.set("Return Product")
            elif barcode == CONSTANTS["cancel"]:
                self.cancel_action()

            elif user := self.database.find_user_by_code(barcode):
                # Set the current user if the user exists
                if user[0] == self.current_user and self.current_user is not None:
                    self.save_data()
                    self.entry_barcode.delete(0, tk.END)
                    self.current_user = None
                    self.label_current_user.config(text="Current User: None")
                    # self.error_label.config(text="")
                    return
                elif user[0] != self.current_user and self.current_user is not None:
                    self.save_data()
                self.current_user = user[0]
                # Update the label displaying the current user
                self.label_current_user.config(text=f"Current User: {user[1]} ({self.current_user})")
                # self.error_label.config(text="")

            else:
                if self.current_user is None:
                    # Display an error message if current_user is None
                    self.error_label.config(text="Zeskanuj kod użytkownika przed skanowaniem produktu")
                    self.entry_barcode.delete(0, tk.END)
                    return

                product_id = barcode

                # Adjust quantity based on the selected activity
                quantity = -1 if self.activity_var.get() == "Take Product" else 1

                date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                # Check if a record with the same Product ID already exists
                item_id = self.find_item_by_product_id(product_id)

                product_row = self.database.find_product_by_code(product_id)
                if not product_row:
                    self.error_label.config(text=f'Produkt z kodem "{product_id}" nie istnieje')
                    self.entry_barcode.delete(0, tk.END)
                    return
                current_quantity = int(product_row[2].value) if product_row else 0
                quantity_in_table = int(self.tree.item(item_id, "values")[3]) if item_id else 0
                print(current_quantity, quantity_in_table + quantity)
                if current_quantity + quantity_in_table + quantity < 0:
                    self.error_label.config(text=f'Niewystarczająca ilość produktu "{product_id}"')
                    self.entry_barcode.delete(0, tk.END)
                    return

                if item_id:
                    # Update the quantity and activity in the table
                    self.update_table_record(item_id, self.current_user, product_id, quantity, date)
                else:
                    # Insert a new record if the Product ID does not exist
                    self.insert_table_record(self.current_user, product_id, quantity, date)

            # Clear the input field
            self.entry_barcode.delete(0, tk.END)

            # Set default focus on barcode input
            self.entry_barcode.focus_set()

    def update_table_record(self, item_id, user_id, product_id, quantity, date):
        current_quantity = int(self.tree.item(item_id, "values")[3])
        updated_quantity = current_quantity + quantity

        # Update the quantity, activity, and user ID in the table
        self.tree.item(item_id, values=(item_id, user_id, product_id, updated_quantity,
                                        "Take Product" if updated_quantity < 0 else "Return Product", date))

        # Remove the record if the quantity becomes zero
        if updated_quantity == 0:
            self.tree.delete(item_id)

    def insert_table_record(self, user_id, product_id, quantity, date):
        # Check if the product code already exists in the 'Products' sheet
        product_row = self.database.find_product_by_code(product_id)

        if product_row:
            # Product code already exists, update the quantity in 'Products' sheet
            # self.error_label.config(text="")
            current_quantity = int(product_row[2].value)
            updated_quantity = current_quantity + int(quantity)

            # Insert a new record in the table
            item_id = str(self.record_id)
            actual_quantity = int(quantity) if updated_quantity >= 0 else -int(quantity)
            self.tree.insert("", "end", iid=item_id, values=(item_id, user_id, product_id, actual_quantity,
                                                             "Take Product" if quantity < 0 else "Return Product",
                                                             date))
            self.record_id += 1  # Increment the record ID

        else:
            # Product code does not exist, display an error message second time just in case XD
            self.error_label.config(text=f'Produkt z kodem "{product_id}" nie istnieje')

    def save_data(self):
        # Iterate over the items in the table and save data to the 'Activities' sheet
        for item_id in self.tree.get_children():
            values = self.tree.item(item_id, "values")
            user_id, product_id, quantity, activity, date = values[1], values[2], values[3], values[4], values[5]
            self.database.add_activity(user_id, product_id, int(quantity), activity, date)

        # Save the workbook
        self.database.workbook.save(self.database.excel_file)

        # Update the quantity of products in the 'Products' sheet
        for item_id in self.tree.get_children():
            values = self.tree.item(item_id, "values")
            product_id, quantity = values[2], values[3]

            # Find the product in the 'Products' sheet
            product_row = self.database.find_product_by_code(product_id)
            if product_row:
                current_quantity = int(product_row[2].value)
                updated_quantity = current_quantity + int(quantity)
                self.database.products.cell(row=product_row[0].row, column=3, value=updated_quantity)

        # Save the workbook again after updating the 'Products' sheet
        self.database.workbook.save(self.database.excel_file)

        # Clear the table
        for item_id in self.tree.get_children():
            self.tree.delete(item_id)

    def find_item_by_product_id(self, product_id):
        for item_id in self.tree.get_children():
            if self.tree.item(item_id, "values")[2] == product_id:
                return item_id
        return None


if __name__ == "__main__":
    excel_file = "inventory.xlsx"
    database = InventoryDatabase(excel_file)

    root = tk.Tk()
    app = InventoryApp(root, database)
    root.mainloop()
